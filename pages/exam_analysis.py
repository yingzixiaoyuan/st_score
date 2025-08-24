"""
考试分析页面模块
处理考试成绩分析和可视化功能
"""

import streamlit as st
import pandas as pd
import plotly.express as px
import plotly.graph_objects as go
from datetime import datetime
import io
from .color_settings import get_score_color, load_color_settings
import openpyxl
from openpyxl.styles import Font, PatternFill


def show_exam_analysis_page(analyzer, exams_df):
    """显示考试分析页面"""
    st.header("📝 考试分析")

    if not exams_df.empty:
        # 考试选择
        st.subheader("📝 考试选择")
        selected_exams = st.multiselect(
            "选择要分析的考试",
            options=exams_df['exam_name'].tolist(),
            default=exams_df['exam_name'].tolist()[:2] if len(
                exams_df) >= 2 else exams_df['exam_name'].tolist(),
            help="默认选择最近的两场考试"
        )

        # 考试概览
        st.subheader("📊 考试概览")
        col1, col2, col3, col4 = st.columns(4)

        with col1:
            st.metric("总考试数", len(exams_df))

        with col2:
            total_students = exams_df['student_count'].sum()
            st.metric("总学生数", total_students)

        with col3:
            avg_students = exams_df['student_count'].mean()
            st.metric("平均参与人数", f"{avg_students:.1f}")

        with col4:
            latest_exam = exams_df.iloc[0]['exam_name']
            st.metric("最新考试", latest_exam)

        # 考试列表
        st.subheader("📋 考试列表")
        st.dataframe(
            exams_df,
            use_container_width=True,
            hide_index=True
        )

        # 成绩分析
        if selected_exams:
            st.header("📈 成绩分析")

            # 获取学生成绩数据
            student_scores = analyzer.get_student_scores(selected_exams)

            if not student_scores.empty:
                # 显示成绩表格
                st.subheader("📊 成绩详情")

                # 操作提示
                st.info("💡 **操作提示**：在下方表格中可以多选学生行（按住Ctrl/Cmd键多选），然后查看选中学生的成绩趋势对比图")

                # 加载颜色设置
                color_settings = load_color_settings()

                # 为每行添加背景颜色（根据平均分）
                def highlight_row_by_average(row):
                    try:
                        avg_score = float(row['平均分'])
                        bg_color = get_score_color(avg_score, color_settings)
                        return [f'background-color: {bg_color}' for _ in row]
                    except Exception:
                        return ['background-color: #F0F0F0' for _ in row]

                # 应用样式并显示表格
                styled_scores = student_scores.style.apply(
                    highlight_row_by_average, axis=1
                ).format({
                    col: "{:.1f}" for col in student_scores.columns
                    if student_scores[col].dtype in ['float64', 'float32'] and col != 'student_id'
                })

                # 使用可选择的表格
                event = st.dataframe(
                    styled_scores,
                    use_container_width=True,
                    hide_index=True,
                    on_select="rerun",
                    selection_mode="multi-row"
                )

                # 学生成绩折线图
                if event and hasattr(event, 'selection') and len(event.selection.rows) > 0:
                    selected_indices = event.selection.rows
                    selected_students = student_scores.iloc[selected_indices]

                    if len(selected_indices) == 1:
                        st.subheader(
                            f"📈 {selected_students.iloc[0]['name']} 的成绩趋势")
                    else:
                        st.subheader(f"📈 学生成绩趋势对比 ({len(selected_indices)}人)")

                    # 获取考试列
                    exam_columns = [
                        col for col in student_scores.columns
                        if col not in ['student_id', 'name', '平均分', '趋势', '等级']
                    ]

                    # 检查是否有足够的考试数据
                    valid_students = []
                    for idx, student in selected_students.iterrows():
                        student_scores_data = []
                        for exam in exam_columns:
                            if pd.notna(student[exam]):
                                student_scores_data.append(student[exam])
                        if len(student_scores_data) >= 2:
                            valid_students.append(student)

                    if len(valid_students) > 0:
                        # 创建多学生对比折线图
                        fig_line = go.Figure()

                        # 定义颜色列表
                        colors = ['#1f77b4', '#ff7f0e', '#2ca02c', '#d62728', '#9467bd',
                                  '#8c564b', '#e377c2', '#7f7f7f', '#bcbd22', '#17becf']

                        for i, student in enumerate(valid_students):
                            exam_scores = []
                            exam_names = []
                            for exam in exam_columns:
                                if pd.notna(student[exam]):
                                    exam_scores.append(student[exam])
                                    exam_names.append(exam)

                            color = colors[i % len(colors)]

                            fig_line.add_trace(go.Scatter(
                                x=exam_names,
                                y=exam_scores,
                                mode='lines+markers',
                                name=f"{student['name']} (平均分: {student['平均分']:.1f})",
                                line=dict(width=3, color=color),
                                marker=dict(size=8, color=color),
                                hovertemplate=f"<b>{student['name']}</b><br>%{{x}}: %{{y:.1f}}分<extra></extra>"
                            ))

                        # 如果只有一个学生，添加平均线和分数标注
                        if len(valid_students) == 1:
                            student = valid_students[0]
                            avg_score = student['平均分']
                            fig_line.add_hline(
                                y=avg_score,
                                line_dash="dash",
                                line_color="red",
                                annotation_text=f"平均分: {avg_score:.1f}分"
                            )

                            # 为单个学生显示分数标注
                            exam_scores = []
                            exam_names = []
                            for exam in exam_columns:
                                if pd.notna(student[exam]):
                                    exam_scores.append(student[exam])
                                    exam_names.append(exam)

                            fig_line.add_trace(go.Scatter(
                                x=exam_names,
                                y=exam_scores,
                                mode='text',
                                text=[f"{score:.1f}分" for score in exam_scores],
                                textposition="top center",
                                textfont=dict(size=10, color=colors[0]),
                                showlegend=False,
                                hoverinfo='skip'
                            ))

                        title = f"成绩趋势对比 ({len(valid_students)}人)" if len(
                            valid_students) > 1 else f"{valid_students[0]['name']} (学号: {valid_students[0]['student_id']}) 的成绩变化"

                        fig_line.update_layout(
                            title=title,
                            xaxis_title="考试",
                            yaxis_title="成绩",
                            yaxis=dict(range=[0, 100]),
                            height=400,
                            showlegend=len(valid_students) > 1,
                            legend=dict(
                                orientation="h",
                                yanchor="bottom",
                                y=1.02,
                                xanchor="right",
                                x=1
                            )
                        )

                        st.plotly_chart(fig_line, use_container_width=True)

                        # 显示学生详细信息
                        if len(valid_students) == 1:
                            # 单个学生的详细信息
                            student = valid_students[0]
                            col1, col2, col3, col4 = st.columns(4)
                            with col1:
                                st.metric("学号", student['student_id'])
                            with col2:
                                st.metric("平均分", f"{student['平均分']:.1f}")
                            with col3:
                                st.metric("成绩趋势", student['趋势'])
                            with col4:
                                st.metric("等级", student['等级'])

                            # 单次考试成绩详情
                            st.markdown("#### 📋 各次考试详细成绩")
                            score_detail = []
                            for exam in exam_columns:
                                if pd.notna(student[exam]):
                                    score_detail.append({
                                        "考试": exam,
                                        "成绩": f"{student[exam]:.1f}",
                                        "等级": analyzer.calculate_level(student[exam])
                                    })

                            score_detail_df = pd.DataFrame(score_detail)
                            st.dataframe(
                                score_detail_df, use_container_width=True, hide_index=True)

                        else:
                            # 多个学生的对比信息
                            st.markdown("#### 📊 选中学生信息对比")

                            comparison_data = []
                            for student in valid_students:
                                comparison_data.append({
                                    "学号": student['student_id'],
                                    "姓名": student['name'],
                                    "平均分": f"{student['平均分']:.1f}",
                                    "趋势": student['趋势'],
                                    "等级": student['等级']
                                })

                            comparison_df = pd.DataFrame(comparison_data)
                            st.dataframe(
                                comparison_df, use_container_width=True, hide_index=True)

                            # 显示统计汇总
                            avg_scores = [student['平均分']
                                          for student in valid_students]
                            col1, col2, col3, col4 = st.columns(4)
                            with col1:
                                st.metric("选中人数", len(valid_students))
                            with col2:
                                st.metric("平均分最高", f"{max(avg_scores):.1f}")
                            with col3:
                                st.metric("平均分最低", f"{min(avg_scores):.1f}")
                            with col4:
                                st.metric(
                                    "组平均分", f"{sum(avg_scores)/len(avg_scores):.1f}")

                    else:
                        # 所有选中的学生数据都不足
                        invalid_names = [student['name']
                                         for _, student in selected_students.iterrows()]
                        st.warning(
                            f"⚠️ 选中的学生 ({', '.join(invalid_names)}) 考试数据不足，无法显示趋势图（需要至少2次考试成绩）")

                else:
                    st.markdown("👆 **请在上方表格中选择学生（可多选），查看成绩趋势图**")
                    st.markdown("💡 **多选提示**：按住 Ctrl/Cmd 键点击可以选择多个学生进行对比分析")

                # 显示颜色说明
                st.markdown("**颜色说明（根据平均分）：**")
                color_legend = []
                for level, config in color_settings.items():
                    color_legend.append(
                        f'<span style="background-color: {config["color"]}; '
                        f'padding: 2px 8px; border-radius: 3px; margin: 2px; '
                        f'display: inline-block;">{level} '
                        f'({config["min_score"]}-{config["max_score"]}分)</span>'
                    )
                st.markdown(" ".join(color_legend), unsafe_allow_html=True)

                # 统计信息
                st.subheader("📊 统计信息")

                # 获取考试列
                score_columns = [
                    col for col in student_scores.columns
                    if col not in ['student_id', 'name', '平均分', '趋势', '等级']
                ]

                col1, col2 = st.columns(2)

                with col1:
                    # 等级分布 - 使用平均分显示（基于用户已选择的考试）
                    level_counts = student_scores['等级'].value_counts()

                    # 根据选择的考试数量确定标题
                    if len(selected_exams) == 1:
                        title = f"成绩等级分布（{selected_exams[0]}）"
                    else:
                        title = f"成绩等级分布（基于{len(selected_exams)}场考试平均分）"

                    # 定义等级颜色
                    level_colors = {
                        '优秀': '#2E8B57',    # 海绿色
                        '良好': '#4169E1',    # 皇家蓝
                        '中等': '#FF8C00',    # 暗橙色
                        '及格': '#FFD700',    # 金色
                        '不及格': '#DC143C'   # 深红色
                    }

                    colors = [level_colors.get(level, '#808080')
                              for level in level_counts.index]

                    fig_level = px.pie(
                        values=level_counts.values,
                        names=level_counts.index,
                        title=title,
                        color_discrete_sequence=colors
                    )
                    st.plotly_chart(fig_level, use_container_width=True)

                with col2:
                    # 趋势分布 - 基于用户已选择的考试
                    trend_counts = student_scores['趋势'].value_counts()

                    # 根据选择的考试数量确定标题和内容
                    if len(selected_exams) == 1:
                        title = "无法显示趋势（仅有1场考试）"
                        # 对于单场考试，可以显示分数分布
                        st.markdown("**📊 单场考试无法显示趋势**")
                        st.info("需要至少2场考试才能分析成绩趋势")

                        # 显示该考试的分数分段分布
                        exam_scores = student_scores[selected_exams[0]].dropna(
                        )
                        score_ranges = []
                        for score in exam_scores:
                            if score >= 90:
                                score_ranges.append('优秀(≥90)')
                            elif score >= 80:
                                score_ranges.append('良好(80-89)')
                            elif score >= 70:
                                score_ranges.append('中等(70-79)')
                            elif score >= 60:
                                score_ranges.append('及格(60-69)')
                            else:
                                score_ranges.append('不及格(<60)')

                        range_counts = pd.Series(score_ranges).value_counts()
                        title = f"分数段分布（{selected_exams[0]}）"

                        # 使用分数段颜色
                        range_colors = {
                            '优秀(≥90)': '#2E8B57',
                            '良好(80-89)': '#4169E1',
                            '中等(70-79)': '#FF8C00',
                            '及格(60-69)': '#FFD700',
                            '不及格(<60)': '#DC143C'
                        }

                        colors = [range_colors.get(r, '#808080')
                                  for r in range_counts.index]

                        total_count = int(range_counts.sum()) if hasattr(range_counts, 'sum') else sum(range_counts.values)
                        percent_labels = [f"{val/total_count*100:.1f}%" for val in range_counts.values]
                        text_labels = [f"{val}人 ({pct})" for val, pct in zip(range_counts.values, percent_labels)]

                        fig_trend = px.bar(
                            x=range_counts.index,
                            y=range_counts.values,
                            title=title,
                            color=range_counts.index,
                            color_discrete_map=range_colors,
                            labels={'x': '分数段', 'y': '人数'},
                            text=text_labels
                        )

                        # 添加数值与百分比标签（双行）
                        fig_trend.update_traces(
                            texttemplate='%{text}',
                            textposition='outside',
                            textfont=dict(size=12)
                        )

                        # 直接显示分数段分布图（带数值标签）
                        st.plotly_chart(fig_trend, use_container_width=True)

                    else:
                        title = f"成绩趋势分布（基于{len(selected_exams)}场考试）"

                        # 定义趋势颜色
                        trend_colors = {
                            '上升': '#32CD32',        # 石灰绿
                            '总体上升': '#228B22',    # 森林绿
                            '下降': '#FF4500',        # 橙红色
                            '总体下降': '#DC143C',    # 深红色
                            '持平': '#4682B4',        # 钢蓝色
                            '波动': '#9370DB',        # 中紫色
                            '数据不足': '#808080'     # 灰色
                        }

                        colors = [trend_colors.get(
                            trend, '#808080') for trend in trend_counts.index]

                        total_count = int(trend_counts.sum()) if hasattr(trend_counts, 'sum') else sum(trend_counts.values)
                        percent_labels = [f"{val/total_count*100:.1f}%" for val in trend_counts.values]
                        text_labels = [f"{val}人 ({pct})" for val, pct in zip(trend_counts.values, percent_labels)]

                        fig_trend = px.bar(
                            x=trend_counts.index,
                            y=trend_counts.values,
                            title=title,
                            color=trend_counts.index,
                            color_discrete_map=trend_colors,
                            labels={'x': '趋势', 'y': '人数'},
                            text=text_labels
                        )

                        # 添加数值标签和百分比
                        fig_trend.update_traces(
                            texttemplate='%{text}',
                            textposition='outside',
                            textfont=dict(size=12)
                        )
                        fig_trend.update_layout(showlegend=False)

                        # 直接显示趋势分布图（带数值标签）
                        st.plotly_chart(fig_trend, use_container_width=True)

                # 成绩对比图
                st.subheader("📊 成绩对比")
                # 过滤出考试名称列（排除其他统计列）
                score_columns = [
                    col for col in student_scores.columns
                    if col not in ['student_id', 'name', '平均分', '趋势', '等级']
                ]

                if len(score_columns) > 1:
                    # 图表类型选择
                    chart_type = st.selectbox(
                        "选择图表类型",
                        ["箱线图", "直方图", "小提琴图", "散点图", "分数段对比"],
                        help="不同图表类型提供不同的数据洞察"
                    )

                    if chart_type == "箱线图":
                        st.markdown("**📈 箱线图说明**：显示数据的五数概括（最小值、q1、中位数、q3、最大值）")
                        st.markdown("- **q1（第一四分位数）**：25%的学生分数低于此值")
                        st.markdown("- **q3（第三四分位数）**：75%的学生分数低于此值")
                        st.markdown("- **箱体范围**：包含中间50%学生的成绩区间")

                        fig_comparison = go.Figure()
                        for exam in score_columns:
                            valid_scores = student_scores[exam].dropna()
                            if not valid_scores.empty:
                                fig_comparison.add_trace(go.Box(
                                    y=valid_scores,
                                    name=exam,
                                    boxpoints='outliers'
                                ))
                        fig_comparison.update_layout(
                            title="各次考试成绩分布对比（箱线图）",
                            yaxis_title="成绩",
                            showlegend=True
                        )
                        st.plotly_chart(fig_comparison, use_container_width=True)

                    elif chart_type == "直方图":
                        st.markdown("**📊 直方图说明**：显示成绩分布的频率，更直观地看出成绩集中区间")

                        fig_comparison = go.Figure()
                        colors = ['#1f77b4', '#ff7f0e', '#2ca02c', '#d62728']
                        for i, exam in enumerate(score_columns):
                            valid_scores = student_scores[exam].dropna()
                            if not valid_scores.empty:
                                fig_comparison.add_trace(go.Histogram(
                                    x=valid_scores,
                                    name=exam,
                                    opacity=0.7,
                                    nbinsx=15,
                                    marker_color=colors[i % len(colors)]
                                ))
                        fig_comparison.update_layout(
                            title="各次考试成绩分布对比（直方图）",
                            xaxis_title="成绩",
                            yaxis_title="人数",
                            barmode='overlay'
                        )
                        st.plotly_chart(fig_comparison, use_container_width=True)

                    elif chart_type == "小提琴图":
                        st.markdown("**🎻 小提琴图说明**：结合箱线图和密度图，显示数据分布形状")

                        fig_comparison = go.Figure()
                        for exam in score_columns:
                            valid_scores = student_scores[exam].dropna()
                            if not valid_scores.empty:
                                fig_comparison.add_trace(go.Violin(
                                    y=valid_scores,
                                    name=exam,
                                    box_visible=True,
                                    meanline_visible=True
                                ))
                        fig_comparison.update_layout(
                            title="各次考试成绩分布对比（小提琴图）",
                            yaxis_title="成绩",
                            showlegend=True
                        )
                        st.plotly_chart(fig_comparison, use_container_width=True)

                    elif chart_type == "散点图" and len(score_columns) >= 2:
                        st.markdown("**📈 散点图说明**：显示两次考试成绩的相关性，每个点代表一个学生")

                        exam1, exam2 = score_columns[0], score_columns[1]
                        # 创建散点图数据，只包含两次考试都有成绩的学生
                        valid_data = student_scores[[
                            exam1, exam2, 'name']].dropna()

                        fig_comparison = go.Figure()
                        fig_comparison.add_trace(go.Scatter(
                            x=valid_data[exam1],
                            y=valid_data[exam2],
                            mode='markers+text',
                            text=valid_data['name'],
                            textposition="top center",
                            marker=dict(size=8, opacity=0.7),
                            name="学生成绩"
                        ))

                        # 添加参考线（x=y）
                        min_score = min(
                            valid_data[exam1].min(), valid_data[exam2].min())
                        max_score = max(
                            valid_data[exam1].max(), valid_data[exam2].max())
                        fig_comparison.add_trace(go.Scatter(
                            x=[min_score, max_score],
                            y=[min_score, max_score],
                            mode='lines',
                            line=dict(dash='dash', color='red'),
                            name="相等线"
                        ))

                        fig_comparison.update_layout(
                            title=f"{exam1} vs {exam2} 成绩对比",
                            xaxis_title=exam1,
                            yaxis_title=exam2
                        )
                        st.plotly_chart(fig_comparison, use_container_width=True)

                    elif chart_type == "分数段对比":
                        st.markdown("**📊 分数段对比说明**：按优秀、良好、中等、及格、不及格分段统计人数")

                        # 定义分数段
                        def get_score_range(score):
                            if score >= 90:
                                return "优秀(90-100)"
                            elif score >= 80:
                                return "良好(80-89)"
                            elif score >= 70:
                                return "中等(70-79)"
                            elif score >= 60:
                                return "及格(60-69)"
                            else:
                                return "不及格(<60)"

                        # 统计各考试的分数段分布
                        range_data = []
                        for exam in score_columns:
                            valid_scores = student_scores[exam].dropna()
                            for score in valid_scores:
                                range_data.append({
                                    'exam': exam,
                                    'range': get_score_range(score),
                                    'score': score
                                })

                        range_df = pd.DataFrame(range_data)
                        range_counts = range_df.groupby(
                            ['exam', 'range']).size().reset_index(name='count')

                        fig_comparison = px.bar(
                            range_counts,
                            x='range',
                            y='count',
                            color='exam',
                            title="各次考试分数段分布对比",
                            labels={'count': '人数', 'range': '分数段'},
                            barmode='group'
                        )

                        # 为对比图添加数值标签
                        fig_comparison.update_traces(
                            texttemplate='%{y}',
                            textposition='outside'
                        )

                        # 直接显示对比图表（带数值标签）
                        st.plotly_chart(
                            fig_comparison, use_container_width=True)

                # 导出功能
                st.subheader("💾 导出结果")
                if st.button("📥 导出到Excel"):
                    # 准备导出数据
                    export_data = student_scores.copy()
                    # 重新排列列顺序，将学号放在最前面
                    column_order = ['student_id', 'name'] + [
                        col for col in export_data.columns if col not in ['student_id', 'name']]
                    export_data = export_data[column_order]
                    # 重命名列名为中文
                    export_data = export_data.rename(columns={
                        'student_id': '学号',
                        'name': '姓名'
                    })

                    # 创建下载链接
                    output = io.BytesIO()

                    # 使用openpyxl引擎，支持样式设置
                    with pd.ExcelWriter(output, engine='openpyxl') as writer:
                        # 先导出数据
                        export_data.to_excel(
                            writer, sheet_name='成绩分析', index=False
                        )

                        # 获取工作表对象
                        worksheet = writer.sheets['成绩分析']

                        # 为每行添加背景颜色
                        for row_idx, (_, row) in enumerate(
                            export_data.iterrows(), start=2
                        ):  # Excel行从2开始（第1行是标题）
                            try:
                                avg_score = float(row['平均分'])
                                bg_color = get_score_color(
                                    avg_score, color_settings)

                                # 将十六进制颜色转换为RGB
                                if bg_color.startswith('#'):
                                    r = int(bg_color[1:3], 16)
                                    g = int(bg_color[3:5], 16)
                                    b = int(bg_color[5:7], 16)

                                    # 为整行设置背景颜色
                                    for col_idx in range(1, len(export_data.columns) + 1):
                                        cell = worksheet.cell(
                                            row=row_idx, column=col_idx)
                                        cell.fill = openpyxl.styles.PatternFill(
                                            start_color=f"{r:02X}{g:02X}{b:02X}",
                                            end_color=f"{r:02X}{g:02X}{b:02X}",
                                            fill_type="solid"
                                        )
                            except Exception:
                                # 如果出错，使用默认颜色
                                pass

                        # 设置标题行样式
                        header_fill = PatternFill(
                            start_color="366092",
                            end_color="366092",
                            fill_type="solid"
                        )
                        header_font = Font(color="FFFFFF", bold=True)

                        for col_idx in range(1, len(export_data.columns) + 1):
                            cell = worksheet.cell(row=1, column=col_idx)
                            cell.fill = header_fill
                            cell.font = header_font

                    output.seek(0)
                    st.download_button(
                        label="📥 下载Excel文件",
                        data=output.getvalue(),
                        file_name=(
                            f"学生成绩分析_"
                            f"{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
                        ),
                        mime=(
                            "application/vnd.openxmlformats-"
                            "officedocument.spreadsheetml.sheet"
                        )
                    )

                    st.success("✅ Excel文件已生成，包含颜色信息！")
            else:
                st.warning("没有找到选中考试的成绩数据")
        else:
            st.info("请在考试选择中选择要分析的考试")
    else:
        st.info("请先导入Excel文件")
